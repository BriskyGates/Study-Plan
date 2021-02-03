import openpyxl
def set_value_by_cell(filename,cell_loc):
    wb=openpyxl.load_workbook(filename)
    ws=wb['Sheet']
    temp=ws[cell_loc].value
    ws['D5']=temp
    wb.save(filename)
    print(temp)
if __name__ == '__main__':
    set_value_by_cell('7_85632465_result.xlsx','B3')